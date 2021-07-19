import os
from collections import Counter
from itertools import takewhile
from openpyxl import Workbook
from openpyxl.utils.exceptions import IllegalCharacterError
from pysubparser import parser
from pysubparser.classes.exceptions import InvalidTimestampError
from pysubparser.cleaners import ascii, brackets, formatting, lower_case
from src.chart_generator import generate_chart
from src.downloader_extractor import sub_download
from src.word_filter import filter_stopwords

# Reads text file which contains IMDB Top 100 movies list.
f = open('seed/movie_list.txt', "r")
subs_to_download = f.readlines()
f.close()

# Downloads subtitles, extracts them.
for subs in subs_to_download:
    sub_download(subs)

# Deletes temporary ".zip" file which is created for download subtitles.
os.remove('subs/temp.zip')

all_subs = []
word_list = ""

# Parse every file with ".srt" extension, in "subs" folder then append them to a list.
[all_subs.append(parser.parse(f'subs/{strs}')) for strs in os.listdir("subs") if strs.endswith(".srt")]

# For every sub file in list, clean subtitles. Adds every line to a list.
for movie_sub in all_subs:
    try:
        movie_sub = ascii.clean(
            brackets.clean(
                lower_case.clean(
                    formatting.clean(movie_sub))))
        for sub_line in movie_sub:
            word_list += sub_line.text
    except UnicodeDecodeError:
        print("Oops something went wrong while decoding this line.")
    except InvalidTimestampError:
        print("Invalid timestamp detected.")

# Convert unwanted signs to space.
for char in '-.,\n?':
    word_list = word_list.replace(char, ' ')

# Split every word as a list element.
words = word_list.split()

# Remove punctuations in words.
for j in range(len(words)):
    words[j] = words[j].replace("'", '')

# Filter some words (e.g. his, she's, the, a, etc.)
words = filter_stopwords(words)

# Keep words more than 10, eliminate the rest.
sorted_words = list(takewhile(lambda i: i[1] > 10, Counter(words).most_common()))

# Export data to the Excel file.
try:
    wbook = Workbook()
    wsheet = wbook.active
    wsheet.append(['Word', 'Count'])
    for x in range(1, len(sorted_words)):
        wsheet.cell(row=x+1, column=1).value = (sorted_words[x-1][0]).encode("ascii", errors="ignore")
    for y in range(1, len(sorted_words)):
        wsheet.cell(row=y+1, column=2).value = sorted_words[y-1][1]
    wbook.save("data_output/Top.Words.List.xlsx")
    print('Excel file saved in the "data_output" folder.')
except IllegalCharacterError:
    print("Illegal char detected.")

# Generate chart of data.
generate_chart("Top.Words.List", "Horizontal")
generate_chart("Top.Words.List", "Vertical")
