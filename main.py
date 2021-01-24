import os
from collections import Counter
from itertools import takewhile
from openpyxl import Workbook
from openpyxl.utils.exceptions import IllegalCharacterError
from pysubparser import parser
from pysubparser.classes.exceptions import InvalidTimestampError
from pysubparser.cleaners import ascii, brackets, formatting, lower_case
from downloader_extractor import sub_download

f = open('./imdb_top_100.txt', "r")
subs_to_download = f.readlines()
f.close()

for subs in subs_to_download:
    sub_download(subs)

os.remove('./subs/temp.zip')

all_subs = []
word_list = ""

[all_subs.append(parser.parse(f'./subs/{strs}')) for strs in os.listdir("./subs") if strs.endswith(".srt")]

for movie_sub in all_subs:
    try:
        sub = movie_sub
        sub = ascii.clean(
            brackets.clean(
                lower_case.clean(
                    formatting.clean(sub))))
        for sub_line in sub:
            word_list += sub_line.text
    except UnicodeDecodeError:
        print("Oops something went wrong while decoding this line.")
    except InvalidTimestampError:
        print("Invalid timestamp detected.")

for char in '-.,\n?':
    word_list = word_list.replace(char, ' ')

words = word_list.split()

# Keep words more than 10, eleminate the rest
sorted_words = list(takewhile(lambda i: i[1] > 10, Counter(words).most_common()))

# Excel
try:
    wbook = Workbook()
    wsheet = wbook.active
    wsheet.append(['Word', 'Count'])
    for x in range(1, len(sorted_words)):
        wsheet.cell(row=x+1, column=1).value = (sorted_words[x-1][0]).encode("ascii", errors="ignore")
    for y in range(1, len(sorted_words)):
        wsheet.cell(row=y+1, column=2).value = sorted_words[y-1][1]
    wbook.save("Top.Words.List.xlsx")
except IllegalCharacterError:
    print("Illegal char detected.")
